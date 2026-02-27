# backend/app/tasks/report_tasks.py
"""Celery tasks for heavy PDF and Excel report generation.
Uses pymongo (sync) — never motor/async in Celery tasks.
Results stored in GridFS; endpoints poll by task_id then download by file_id.
"""

import io
from datetime import datetime, timedelta
from pymongo import MongoClient
from celery import shared_task
from app.config import settings
from app.services.gridfs_service import sync_gridfs_service


# ── helpers ──────────────────────────────────────────────────────────────────

def _get_db():
    client = MongoClient(settings.MONGODB_URL)
    return client, client[settings.MONGODB_DB_NAME]


def _farmer_name(farmer: dict) -> str:
    pi = farmer.get("personal_info", {})
    return f"{pi.get('first_name', '')} {pi.get('last_name', '')}".strip() or "Unknown"


def _fmt_date(dt) -> str:
    if not dt:
        return "N/A"
    if isinstance(dt, datetime):
        return dt.strftime("%d %b %Y")
    return str(dt)


# ── PDF helpers ───────────────────────────────────────────────────────────────

def _pdf_table_style():
    from reportlab.lib import colors
    from reportlab.platypus import TableStyle
    return TableStyle([
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.HexColor("#f9fafb"), colors.HexColor("#ffffff")]),
        ("GRID",           (0, 0), (-1, -1), 0.4, colors.HexColor("#e5e7eb")),
        ("FONTNAME",       (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE",       (0, 0), (-1, -1), 8),
        ("PADDING",        (0, 0), (-1, -1), 5),
        ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
        ("TEXTCOLOR",      (0, 0), (-1, -1), colors.HexColor("#374151")),
    ])


def _header_style():
    from reportlab.lib import colors
    from reportlab.platypus import TableStyle
    return TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#15803d")),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0), 9),
        ("PADDING",     (0, 0), (-1, -1), 6),
        ("GRID",        (0, 0), (-1, -1), 0.4, colors.HexColor("#e5e7eb")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#f9fafb"), colors.HexColor("#ffffff")]),
        ("FONTNAME",    (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",    (0, 1), (-1, -1), 8),
        ("TEXTCOLOR",   (0, 1), (-1, -1), colors.HexColor("#374151")),
    ])


def _build_pdf_header(story, title: str, subtitle: str = ""):
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.platypus import Paragraph, Spacer
    from reportlab.lib.units import cm

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("CemTitle", parent=styles["Heading1"],
                                  fontSize=16, textColor=colors.HexColor("#15803d"),
                                  alignment=TA_CENTER, spaceAfter=4)
    sub_style   = ParagraphStyle("CemSub",   parent=styles["Normal"],
                                  fontSize=10, textColor=colors.HexColor("#6b7280"),
                                  alignment=TA_CENTER)
    story.append(Paragraph("Zambian Farmer Registration System — CEM", title_style))
    story.append(Paragraph(title, styles["Heading2"]))
    if subtitle:
        story.append(Paragraph(subtitle, sub_style))
    story.append(Spacer(1, 0.4 * cm))
    return styles


# ── TASK 1: Farmer Profile PDF ────────────────────────────────────────────────

@shared_task(name="app.tasks.report_tasks.generate_farmer_pdf", bind=True)
def generate_farmer_pdf(self, farmer_id: str) -> dict:
    """Generate full farmer profile PDF; store in GridFS; return file_id."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Spacer, Table, Paragraph, Image, HRFlowable
    )

    client, db = _get_db()
    try:
        farmer = db.farmers.find_one({"farmer_id": farmer_id})
        if not farmer:
            raise ValueError(f"Farmer {farmer_id} not found")

        operator = None
        if farmer.get("created_by"):
            operator = db.operators.find_one(
                {"email": farmer["created_by"]},
                {"full_name": 1, "operator_id": 1, "email": 1}
            )

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
                                 rightMargin=2 * cm, leftMargin=2 * cm,
                                 topMargin=2 * cm, bottomMargin=2 * cm)
        story = []
        styles = _build_pdf_header(story, "Farmer Profile Report",
                                    f"Generated: {datetime.utcnow().strftime('%d %B %Y %H:%M UTC')}")

        pi   = farmer.get("personal_info", {})
        addr = farmer.get("address", {})
        farm = farmer.get("farm_info", {})

        # ── Personal details ──────────────────────────────────────────────────
        story.append(Paragraph("Personal Information", styles["Heading3"]))
        personal = [
            ["Farmer ID",   farmer.get("farmer_id", "N/A"),
             "Full Name",   f"{pi.get('first_name','')} {pi.get('last_name','')}".strip()],
            ["NRC",         pi.get("nrc", "N/A"),
             "Gender",      pi.get("gender", "N/A")],
            ["Date of Birth", _fmt_date(pi.get("date_of_birth")),
             "Ethnic Group", pi.get("ethnic_group", "N/A")],
            ["Phone (Primary)", pi.get("phone_primary", "N/A"),
             "Phone (Secondary)", pi.get("phone_secondary", "N/A")],
            ["Email",       pi.get("email", "N/A"),
             "Registered",  _fmt_date(farmer.get("created_at"))],
        ]
        t = Table(personal, colWidths=[3.5*cm, 5*cm, 3.5*cm, 5*cm])
        ts = _pdf_table_style()
        ts.add("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold")
        ts.add("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold")
        t.setStyle(ts)
        story.append(t)
        story.append(Spacer(1, 0.4 * cm))

        # ── Address ───────────────────────────────────────────────────────────
        story.append(Paragraph("Address", styles["Heading3"]))
        address_data = [
            ["Province", addr.get("province_name", "N/A"),
             "District",  addr.get("district_name", "N/A")],
            ["Chiefdom",  addr.get("chiefdom_name", "N/A"),
             "Village",   addr.get("village", "N/A")],
        ]
        t2 = Table(address_data, colWidths=[3.5*cm, 5*cm, 3.5*cm, 5*cm])
        ts2 = _pdf_table_style()
        ts2.add("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold")
        ts2.add("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold")
        t2.setStyle(ts2)
        story.append(t2)
        story.append(Spacer(1, 0.4 * cm))

        # ── Farm info ─────────────────────────────────────────────────────────
        story.append(Paragraph("Farm Information", styles["Heading3"]))
        crops     = ", ".join(farm.get("crops_grown", [])) or "N/A"
        livestock = ", ".join(farm.get("livestock_types", farm.get("livestock", []))) or "N/A"
        farm_data = [
            ["Farm Size (ha)", str(farm.get("farm_size_hectares", "N/A")),
             "Experience (yrs)", str(farm.get("farming_experience_years", farm.get("years_farming", "N/A")))],
            ["Crops Grown",   crops,
             "Irrigation",    "Yes" if farm.get("has_irrigation") else "No"],
            ["Livestock",     livestock,
             "Household Size", str(farmer.get("household_size", farm.get("household_size", "N/A")))],
        ]
        t3 = Table(farm_data, colWidths=[3.5*cm, 5*cm, 3.5*cm, 5*cm])
        ts3 = _pdf_table_style()
        ts3.add("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold")
        ts3.add("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold")
        t3.setStyle(ts3)
        story.append(t3)
        story.append(Spacer(1, 0.4 * cm))

        # ── Operator mapping ──────────────────────────────────────────────────
        if operator:
            story.append(Paragraph("Registered By (Operator)", styles["Heading3"]))
            op_data = [
                ["Operator ID",   operator.get("operator_id", "N/A"),
                 "Name",          operator.get("full_name", "N/A")],
                ["Email",         operator.get("email", "N/A"), "", ""],
            ]
            t4 = Table(op_data, colWidths=[3.5*cm, 5*cm, 3.5*cm, 5*cm])
            ts4 = _pdf_table_style()
            ts4.add("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold")
            ts4.add("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold")
            t4.setStyle(ts4)
            story.append(t4)
            story.append(Spacer(1, 0.4 * cm))

        # ── Photo from GridFS ─────────────────────────────────────────────────
        photo_file_id = (
            farmer.get("photo_file_id")
            or (farmer.get("documents") or {}).get("photo_file_id")
        )
        if photo_file_id:
            try:
                photo_bytes, _ = sync_gridfs_service.download_file(str(photo_file_id))
                story.append(Paragraph("Farmer Photo", styles["Heading3"]))
                img = Image(io.BytesIO(photo_bytes), width=4 * cm, height=4 * cm)
                story.append(img)
                story.append(Spacer(1, 0.3 * cm))
            except Exception:
                pass  # Photo unavailable — skip silently

        # ── Documents list ────────────────────────────────────────────────────
        docs = farmer.get("documents", {})
        if docs:
            story.append(Paragraph("Documents on File", styles["Heading3"]))
            doc_keys = [k for k in docs if k not in ("photo_file_id", "photoFileId")]
            if doc_keys:
                doc_rows = [[k, str(docs[k])[:60]] for k in doc_keys]
                td = Table([["Document", "Reference"]] + doc_rows, colWidths=[6 * cm, 11 * cm])
                td.setStyle(_header_style())
                story.append(td)
            story.append(Spacer(1, 0.4 * cm))

        # ── Footer ────────────────────────────────────────────────────────────
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#e5e7eb")))
        story.append(Spacer(1, 0.2 * cm))
        story.append(Paragraph(
            f"CEM Platform • Confidential • Generated {datetime.utcnow().strftime('%d %B %Y %H:%M UTC')}",
            styles["Normal"]
        ))

        doc.build(story)
        pdf_bytes = buf.getvalue()

        file_id = sync_gridfs_service.upload_file(
            file_data=pdf_bytes,
            filename=f"{farmer_id}_profile.pdf",
            farmer_id=farmer_id,
            file_type="report",
        )
        return {"status": "completed", "file_id": file_id, "filename": f"{farmer_id}_profile.pdf"}
    finally:
        client.close()


# ── TASK 2: Operator Report PDF ───────────────────────────────────────────────

@shared_task(name="app.tasks.report_tasks.generate_operator_pdf", bind=True)
def generate_operator_pdf(self, operator_id: str) -> dict:
    """Generate operator report PDF; return file_id."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Spacer, Table, Paragraph, HRFlowable

    client, db = _get_db()
    try:
        operator = db.operators.find_one({"operator_id": operator_id})
        if not operator:
            raise ValueError(f"Operator {operator_id} not found")

        farmers     = list(db.farmers.find({"created_by": operator.get("email", "")},
                                           {"farmer_id": 1, "personal_info": 1, "address": 1, "created_at": 1}))
        cutoff      = datetime.utcnow() - timedelta(days=30)
        recent_count = sum(1 for f in farmers if (f.get("created_at") or datetime.min) >= cutoff)

        buf  = io.BytesIO()
        doc  = SimpleDocTemplate(buf, pagesize=A4,
                                  rightMargin=2*cm, leftMargin=2*cm,
                                  topMargin=2*cm,   bottomMargin=2*cm)
        story = []
        styles = _build_pdf_header(story, "Operator Report",
                                    f"Operator ID: {operator_id} • Generated: {datetime.utcnow().strftime('%d %B %Y')}")

        # ── Operator Profile ──────────────────────────────────────────────────
        story.append(Paragraph("Operator Profile", styles["Heading3"]))
        op_data = [
            ["Operator ID",  operator.get("operator_id", "N/A"), "Full Name", operator.get("full_name", "N/A")],
            ["Email",        operator.get("email", "N/A"),       "Phone",     operator.get("phone", "N/A")],
            ["Province",     operator.get("province_name", operator.get("province", "N/A")),
             "District",     operator.get("district_name",  operator.get("district", "N/A"))],
            ["Status",       "Active" if operator.get("is_active", True) else "Inactive",
             "Registered",   _fmt_date(operator.get("created_at"))],
        ]
        t = Table(op_data, colWidths=[3.5*cm, 5*cm, 3.5*cm, 5*cm])
        ts = _pdf_table_style()
        ts.add("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold")
        ts.add("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold")
        t.setStyle(ts)
        story.append(t)
        story.append(Spacer(1, 0.4 * cm))

        # ── Activity Metrics ──────────────────────────────────────────────────
        story.append(Paragraph("Activity Metrics", styles["Heading3"]))
        metrics_data = [
            ["Total Farmers Registered", str(len(farmers))],
            ["Registrations (Last 30 Days)", str(recent_count)],
        ]
        tm = Table(metrics_data, colWidths=[9*cm, 8*cm])
        tsm = _pdf_table_style()
        tsm.add("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold")
        tm.setStyle(tsm)
        story.append(tm)
        story.append(Spacer(1, 0.4 * cm))

        # ── Farmer List ───────────────────────────────────────────────────────
        story.append(Paragraph(f"Registered Farmers ({len(farmers)} total)", styles["Heading3"]))
        farmer_rows = [["#", "Farmer ID", "Name", "Province", "District", "Registered"]]
        for i, f in enumerate(farmers[:200], 1):  # cap at 200 rows in PDF
            farmer_rows.append([
                str(i),
                f.get("farmer_id", "N/A"),
                _farmer_name(f),
                (f.get("address") or {}).get("province_name", "N/A"),
                (f.get("address") or {}).get("district_name", "N/A"),
                _fmt_date(f.get("created_at")),
            ])
        tf = Table(farmer_rows, colWidths=[0.8*cm, 3*cm, 4*cm, 3*cm, 3*cm, 3*cm])
        tf.setStyle(_header_style())
        story.append(tf)

        if len(farmers) > 200:
            story.append(Spacer(1, 0.2*cm))
            story.append(Paragraph(f"... and {len(farmers)-200} more. Download Excel for full list.", styles["Normal"]))

        story.append(Spacer(1, 0.4*cm))
        story.append(HRFlowable(width="100%", thickness=0.5,
                                 color=colors.HexColor("#e5e7eb")))
        story.append(Spacer(1, 0.2*cm))
        story.append(Paragraph(
            f"CEM Platform • Confidential • Generated {datetime.utcnow().strftime('%d %B %Y %H:%M UTC')}",
            styles["Normal"]
        ))

        doc.build(story)
        file_id = sync_gridfs_service.upload_file(
            file_data=buf.getvalue(),
            filename=f"operator_{operator_id}_report.pdf",
            farmer_id=operator_id,
            file_type="report",
        )
        return {"status": "completed", "file_id": file_id,
                "filename": f"operator_{operator_id}_report.pdf"}
    finally:
        client.close()


# ── TASK 3: Admin Summary PDF ─────────────────────────────────────────────────

@shared_task(name="app.tasks.report_tasks.generate_summary_pdf", bind=True)
def generate_summary_pdf(self) -> dict:
    """Generate admin summary PDF; return file_id."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Spacer, Table, Paragraph, HRFlowable

    client, db = _get_db()
    try:
        # Aggregations
        total_farmers   = db.farmers.count_documents({})
        total_operators = db.operators.count_documents({})
        total_users     = db.users.count_documents({})
        month_start     = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        this_month      = db.farmers.count_documents({"created_at": {"$gte": month_start}})

        by_province = list(db.farmers.aggregate([
            {"$group": {"_id": "$address.province_name", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}},
        ]))
        by_operator = list(db.farmers.aggregate([
            {"$group": {"_id": "$created_by", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}},
            {"$limit": 20},
        ]))
        crops_pipeline = list(db.farmers.aggregate([
            {"$unwind": "$farm_info.crops_grown"},
            {"$group": {"_id": "$farm_info.crops_grown", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}},
        ]))

        buf  = io.BytesIO()
        doc  = SimpleDocTemplate(buf, pagesize=A4,
                                  rightMargin=2*cm, leftMargin=2*cm,
                                  topMargin=2*cm,   bottomMargin=2*cm)
        story = []
        styles = _build_pdf_header(story, "Admin Summary Report",
                                    f"Generated: {datetime.utcnow().strftime('%d %B %Y %H:%M UTC')}")

        # ── Totals ────────────────────────────────────────────────────────────
        story.append(Paragraph("System Totals", styles["Heading3"]))
        totals = [
            ["Metric", "Count"],
            ["Total Farmers Registered",  str(total_farmers)],
            ["Total Operators",           str(total_operators)],
            ["Total Users",               str(total_users)],
            ["Farmers Registered This Month", str(this_month)],
        ]
        tt = Table(totals, colWidths=[10*cm, 7*cm])
        tt.setStyle(_header_style())
        story.append(tt)
        story.append(Spacer(1, 0.5*cm))

        # ── By Province ───────────────────────────────────────────────────────
        story.append(Paragraph("Farmers by Province", styles["Heading3"]))
        prov_rows = [["Province", "Farmer Count"]]
        for r in by_province:
            prov_rows.append([r.get("_id") or "Unknown", str(r["count"])])
        tp = Table(prov_rows, colWidths=[10*cm, 7*cm])
        tp.setStyle(_header_style())
        story.append(tp)
        story.append(Spacer(1, 0.5*cm))

        # ── By Operator ───────────────────────────────────────────────────────
        story.append(Paragraph("Farmers by Operator (Top 20)", styles["Heading3"]))
        op_rows = [["Operator Email", "Farmer Count"]]
        for r in by_operator:
            op_rows.append([r.get("_id") or "Unknown", str(r["count"])])
        to = Table(op_rows, colWidths=[10*cm, 7*cm])
        to.setStyle(_header_style())
        story.append(to)
        story.append(Spacer(1, 0.5*cm))

        # ── Crops by Region ───────────────────────────────────────────────────
        if crops_pipeline:
            story.append(Paragraph("Crops Distribution", styles["Heading3"]))
            crop_rows = [["Crop", "Count"]]
            for r in crops_pipeline:
                crop_rows.append([r.get("_id") or "Unknown", str(r["count"])])
            tc = Table(crop_rows, colWidths=[10*cm, 7*cm])
            tc.setStyle(_header_style())
            story.append(tc)
            story.append(Spacer(1, 0.5*cm))

        story.append(HRFlowable(width="100%", thickness=0.5,
                                 color=colors.HexColor("#e5e7eb")))
        story.append(Spacer(1, 0.2*cm))
        story.append(Paragraph(
            f"CEM Platform • Admin Only • Confidential • Generated {datetime.utcnow().strftime('%d %B %Y %H:%M UTC')}",
            styles["Normal"]
        ))

        doc.build(story)
        file_id = sync_gridfs_service.upload_file(
            file_data=buf.getvalue(),
            filename="summary_report.pdf",
            farmer_id="admin",
            file_type="report",
        )
        return {"status": "completed", "file_id": file_id, "filename": "summary_report.pdf"}
    finally:
        client.close()


# ── TASK 4: Farmers Excel Export ──────────────────────────────────────────────

@shared_task(name="app.tasks.report_tasks.generate_farmers_excel", bind=True)
def generate_farmers_excel(self, province: str = None) -> dict:
    """Export all farmers to Excel with clean column headers."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    client, db = _get_db()
    try:
        query = {}
        if province:
            query["address.province_name"] = province

        farmers = list(db.farmers.find(query, {
            "farmer_id": 1, "personal_info": 1, "address": 1,
            "farm_info": 1, "created_at": 1, "created_by": 1, "status": 1,
        }))

        wb = Workbook()
        ws = wb.active
        ws.title = "Farmers"

        headers = [
            "Farmer ID", "First Name", "Last Name", "NRC", "Gender", "Date of Birth",
            "Ethnic Group", "Phone Primary", "Phone Secondary", "Email",
            "Province", "District", "Chiefdom", "Village",
            "Farm Size (ha)", "Crops Grown", "Livestock", "Has Irrigation",
            "Experience (yrs)", "Household Size", "Primary Income Source",
            "Status", "Operator Email", "Registered Date",
        ]

        # Header row style
        header_fill = PatternFill(start_color="15803d", end_color="15803d", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.row_dimensions[1].height = 30

        for row_idx, f in enumerate(farmers, 2):
            pi   = f.get("personal_info", {}) or {}
            addr = f.get("address", {}) or {}
            farm = f.get("farm_info", {}) or {}
            ws.cell(row=row_idx, column=1,  value=f.get("farmer_id", ""))
            ws.cell(row=row_idx, column=2,  value=pi.get("first_name", ""))
            ws.cell(row=row_idx, column=3,  value=pi.get("last_name", ""))
            ws.cell(row=row_idx, column=4,  value=pi.get("nrc", ""))
            ws.cell(row=row_idx, column=5,  value=pi.get("gender", ""))
            ws.cell(row=row_idx, column=6,  value=str(pi.get("date_of_birth", "")))
            ws.cell(row=row_idx, column=7,  value=pi.get("ethnic_group", ""))
            ws.cell(row=row_idx, column=8,  value=pi.get("phone_primary", ""))
            ws.cell(row=row_idx, column=9,  value=pi.get("phone_secondary", ""))
            ws.cell(row=row_idx, column=10, value=pi.get("email", ""))
            ws.cell(row=row_idx, column=11, value=addr.get("province_name", ""))
            ws.cell(row=row_idx, column=12, value=addr.get("district_name", ""))
            ws.cell(row=row_idx, column=13, value=addr.get("chiefdom_name", ""))
            ws.cell(row=row_idx, column=14, value=addr.get("village", ""))
            ws.cell(row=row_idx, column=15, value=farm.get("farm_size_hectares", ""))
            ws.cell(row=row_idx, column=16, value=", ".join(farm.get("crops_grown", [])))
            ws.cell(row=row_idx, column=17,
                    value=", ".join(farm.get("livestock_types", farm.get("livestock", []))))
            ws.cell(row=row_idx, column=18, value="Yes" if farm.get("has_irrigation") else "No")
            ws.cell(row=row_idx, column=19,
                    value=farm.get("farming_experience_years", farm.get("years_farming", "")))
            ws.cell(row=row_idx, column=20,
                    value=str(f.get("household_size", farm.get("household_size", ""))))
            ws.cell(row=row_idx, column=21, value=farm.get("primary_income_source", ""))
            ws.cell(row=row_idx, column=22, value=f.get("status", "active"))
            ws.cell(row=row_idx, column=23, value=f.get("created_by", ""))
            created = f.get("created_at")
            ws.cell(row=row_idx, column=24,
                    value=created.strftime("%Y-%m-%d") if isinstance(created, datetime) else str(created or ""))

        # Auto column widths
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)

        file_id = sync_gridfs_service.upload_file(
            file_data=buf.getvalue(),
            filename="farmers_export.xlsx",
            farmer_id="admin",
            file_type="report",
        )
        return {"status": "completed", "file_id": file_id,
                "filename": "farmers_export.xlsx", "total_rows": len(farmers)}
    finally:
        client.close()


# ── TASK 5: Admin Summary Excel ───────────────────────────────────────────────

@shared_task(name="app.tasks.report_tasks.generate_summary_excel", bind=True)
def generate_summary_excel(self) -> dict:
    """Export admin summary statistics to Excel (multiple sheets)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    client, db = _get_db()
    try:
        total_farmers   = db.farmers.count_documents({})
        total_operators = db.operators.count_documents({})
        total_users     = db.users.count_documents({})
        month_start     = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        this_month      = db.farmers.count_documents({"created_at": {"$gte": month_start}})

        by_province = list(db.farmers.aggregate([
            {"$group": {"_id": "$address.province_name", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}},
        ]))
        by_operator = list(db.farmers.aggregate([
            {"$group": {"_id": "$created_by", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}},
        ]))
        crops_data = list(db.farmers.aggregate([
            {"$unwind": "$farm_info.crops_grown"},
            {"$group": {"_id": "$farm_info.crops_grown", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}},
        ]))

        wb = Workbook()
        hdr_fill = PatternFill(start_color="15803d", end_color="15803d", fill_type="solid")
        hdr_font = Font(color="FFFFFF", bold=True)

        def write_sheet(ws, title_row, headers, rows):
            ws.cell(row=1, column=1, value=title_row).font = Font(bold=True, size=12)
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=2, column=col, value=h)
                c.fill, c.font = hdr_fill, hdr_font
                c.alignment = Alignment(horizontal="center")
            for r_idx, row in enumerate(rows, 3):
                for c_idx, val in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=val)
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(
                    max(len(str(c.value or "")) for c in col) + 4, 40
                )
            ws.freeze_panes = "A3"

        # Sheet 1: Overview
        ws1 = wb.active
        ws1.title = "Overview"
        write_sheet(ws1, "System Overview", ["Metric", "Value"], [
            ["Total Farmers",              total_farmers],
            ["Total Operators",            total_operators],
            ["Total Users",                total_users],
            ["Farmers Registered (Month)", this_month],
            ["Report Generated",           datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")],
        ])

        # Sheet 2: By Province
        ws2 = wb.create_sheet("Farmers by Province")
        write_sheet(ws2, "Farmer Count by Province", ["Province", "Farmer Count"],
                    [[r.get("_id") or "Unknown", r["count"]] for r in by_province])

        # Sheet 3: By Operator
        ws3 = wb.create_sheet("Farmers by Operator")
        write_sheet(ws3, "Farmer Count by Operator", ["Operator Email", "Farmer Count"],
                    [[r.get("_id") or "Unknown", r["count"]] for r in by_operator])

        # Sheet 4: Crops
        if crops_data:
            ws4 = wb.create_sheet("Crops Distribution")
            write_sheet(ws4, "Crop Count", ["Crop", "Count"],
                        [[r.get("_id") or "Unknown", r["count"]] for r in crops_data])

        buf = io.BytesIO()
        wb.save(buf)

        file_id = sync_gridfs_service.upload_file(
            file_data=buf.getvalue(),
            filename="summary_report.xlsx",
            farmer_id="admin",
            file_type="report",
        )
        return {"status": "completed", "file_id": file_id, "filename": "summary_report.xlsx"}
    finally:
        client.close()
